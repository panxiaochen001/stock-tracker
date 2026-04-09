import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RED_FILL    = PatternFill("solid", start_color="FFE6E6")
GREEN_FILL  = PatternFill("solid", start_color="E6FFE6")
HEADER_FILL = PatternFill("solid", start_color="1F3864")
ACTIVE_FILL = PatternFill("solid", start_color="FFF3CD")
RED_FONT    = Font(color="C0392B")
GREEN_FONT  = Font(color="1A7A1A")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=10)
THIN        = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)
PERIODS = ["5日", "10日", "1月", "2月", "3月"]


def _write_header(ws, cols):
    for ci, name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN
    ws.row_dimensions[1].height = 22


def _pct_cell(cell, tup):
    if not isinstance(tup, tuple) or tup[0] is None:
        cell.value = "—"
        return
    val, status = tup
    prefix = "▶ " if status == "进行中" else ""
    cell.value     = f"{prefix}{val:+.2f}%"
    cell.alignment = Alignment(horizontal="center")
    cell.border    = THIN
    if status == "进行中":
        cell.fill = ACTIVE_FILL
    elif val > 0:
        cell.font = RED_FONT
        cell.fill = RED_FILL
    elif val < 0:
        cell.font = GREEN_FONT
        cell.fill = GREEN_FILL


def build_excel(rows: list[dict]) -> bytes:
    wb = Workbook()

    # ── Sheet1 明细 ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = "选股明细"
    ws1.freeze_panes = "A2"

    base_cols = ["股票代码", "股票名称", "选股日", "买入日", "买入价(元)"]
    pct_cols  = [f"{p}涨幅"     for p in PERIODS]
    high_cols = [f"{p}最高涨幅" for p in PERIODS]
    all_cols  = base_cols + pct_cols + high_cols + ["备注"]
    _write_header(ws1, all_cols)

    for ri, row in enumerate(rows, 2):
        ci = 1
        for col in base_cols:
            cell = ws1.cell(row=ri, column=ci, value=row.get(col))
            cell.alignment = Alignment(horizontal="center")
            cell.border    = THIN
            if col == "买入价(元)" and row.get(col):
                cell.number_format = "0.00"
            ci += 1
        for col in pct_cols + high_cols:
            _pct_cell(ws1.cell(row=ri, column=ci), row.get(col, (None, "")))
            ci += 1
        cell = ws1.cell(row=ri, column=ci, value=row.get("备注", ""))
        cell.border = THIN

    widths = [12, 10, 10, 10, 12] + [10] * 10 + [16]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet2 统计 ───────────────────────────────────────
    ws2 = wb.create_sheet("统计汇总")
    ws2.freeze_panes = "A2"

    stat_cols = (["统计维度", "股票数"]
                 + [f"{p}均涨幅" for p in PERIODS]
                 + [f"{p}均最高" for p in PERIODS]
                 + [f"{p}胜率"   for p in PERIODS])
    _write_header(ws2, stat_cols)

    def _stat(label, subset):
        r = {"统计维度": label, "股票数": len(subset)}
        for p in PERIODS:
            vals  = [row[f"{p}涨幅"][0]     for row in subset
                     if isinstance(row.get(f"{p}涨幅"), tuple)
                     and row[f"{p}涨幅"][0] is not None]
            highs = [row[f"{p}最高涨幅"][0] for row in subset
                     if isinstance(row.get(f"{p}最高涨幅"), tuple)
                     and row[f"{p}最高涨幅"][0] is not None]
            r[f"{p}均涨幅"] = round(sum(vals)/len(vals),   2) if vals  else None
            r[f"{p}均最高"] = round(sum(highs)/len(highs), 2) if highs else None
            r[f"{p}胜率"]   = f"{sum(1 for v in vals if v>0)/len(vals)*100:.0f}%" if vals else "—"
        return r

    from itertools import groupby
    stat_rows = [_stat("全部", rows)]
    for dk, grp in groupby(sorted(rows, key=lambda x: x.get("选股日","")),
                            key=lambda x: x.get("选股日","")):
        stat_rows.append(_stat(f"选股日 {dk}", list(grp)))

    for ri, srow in enumerate(stat_rows, 2):
        for ci, col in enumerate(stat_cols, 1):
            val  = srow.get(col)
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = THIN
            if isinstance(val, float):
                cell.font = RED_FONT if val > 0 else (GREEN_FONT if val < 0 else cell.font)
                cell.fill = RED_FILL if val > 0 else (GREEN_FILL if val < 0 else cell.fill)

    for i, w in enumerate([18, 8] + [10]*15, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
